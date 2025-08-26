from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional, Literal, Iterable
from datetime import datetime, timezone
import json, io, os, gzip, re
from openpyxl import load_workbook

SHARED_FOLDER = "/app/shared"
DATA_JSON = os.path.join(SHARED_FOLDER, "data.json")
LISTEN_FOLDER = os.path.join(SHARED_FOLDER, "Expose", "Listen")
PROTOKOLL_FOLDER = os.path.join(SHARED_FOLDER, "Expose", "Protokolle")
META_JSON = os.path.join(SHARED_FOLDER, "meta.json")

def compose_response_structure(
    vn_nr: str,
    _excel_template_path: Optional[str],   # wird durch data.json übersteuert
    _protokoll_root: str,                  # wir nutzen die Konstanten oben
    *,
    output: Literal["json","tsv"] = "tsv"
) -> bytes | Dict[str, Any]:
    """Baut zuerst dein JSON-Modell, optional als TSV ausgegeben."""
    # 1) filename über data.json
    filename = _resolve_filename_from_datajson(vn_nr)
    if not filename:
        raise FileNotFoundError(f"VN {vn_nr}: kein Eintrag in data.json gefunden.")

    listen_path = Path(LISTEN_FOLDER) / filename
    if not listen_path.exists():
        raise FileNotFoundError(f"Listen-Datei fehlt: {listen_path}")

    proto_path = _resolve_protokoll_path(vn_nr, filename)  # kann None sein

    # 2) Listen lesen (alle Sheets = Anlagen)
    listen_struct, meta_from_listen = _parse_listen_workbook(listen_path)

    # 3) Protokoll lesen (optional)
    protokoll_struct = _parse_protokoll_workbook(proto_path) if proto_path else {}

    # 4) Merge je Anlage (Sheetname = Anlagenname)
    anlagen: List[Dict[str, Any]] = []
    for anlagen_name, ldata in listen_struct.items():
        groups_out: List[Dict[str, Any]] = []
        # Map aus Protokoll: (gruppe, melder_index) -> "Q1" / "N.i.o." / ""
        pmap = protokoll_struct.get(anlagen_name, {})

        for g in ldata["Gruppen"]:
            mg = str(g["Nummer"])
            anz = g["Anz"]
            art = g["Art"]
            melder = []
            for idx in range(1, anz + 1):
                typ = g["MelderTypen"][idx-1] if idx-1 < len(g["MelderTypen"]) and g["MelderTypen"][idx-1] else "-"
                ausl = pmap.get((mg, idx), "-")
                melder.append({"Nummer": idx, "typ": typ, "ausluesung": ausl or "-"})
            groups_out.append({"Nummer": int(mg) if mg.isdigit() else mg, "Art": art or "-", "Melder": melder})

        anlagen.append({
            "name": anlagen_name,
            "Gruppen": groups_out,
            "Hardware": ldata["Hardware"],  # aus Vorlage; falls nicht vorhanden: leer
            "Wartung": ldata["Wartung"],    # aus Vorlage; sonst Q1..Q4-Defaults
        })

    # 5) Meta (Kunde/Wartungstyp) aus data.json oder aus Listen-Kopf
    kunde, wartungstyp = _resolve_meta(vn_nr, meta_from_listen)

    model = {
        "Data": {
            "Meta": {
                "VN": _normalize_vn(vn_nr),
                "Kunde": kunde or "-",
                "Wartungstyp": wartungstyp or "-",
                "Anlagen": anlagen
            }
        }
    }

    if output == "json":
        return model
    return _pack_tsv(model)

# ----------------------------- Resolver -------------------------------------

def _resolve_filename_from_datajson(vn_nr: str) -> Optional[str]:
    norm_vn = _normalize_vn(vn_nr)
    if not os.path.exists(DATA_JSON):
        return None
    with open(DATA_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)
    for entry in data.get("vn", []):
        if _normalize_vn(entry.get("vn_nr","")) == norm_vn:
            return entry.get("filename")
    return None

def _resolve_protokoll_path(vn_nr: str, filename: str) -> Optional[Path]:
    vn_dir = Path(PROTOKOLL_FOLDER) / _normalize_vn(vn_nr)
    exact = vn_dir / filename
    if exact.exists():
        return exact
    if vn_dir.exists():
        files = sorted(vn_dir.rglob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        return files[0] if files else None
    return None

def _normalize_vn(vn: str) -> str:
    s = (vn or "").strip().upper()
    return s if s.startswith("VN") else f"VN{s}"

# ----------------------------- Parser (fest auf eure Struktur) --------------

def _parse_listen_workbook(path: Path) -> Tuple[Dict[str, Any], Dict[str, str]]:
    """
    Pro Sheet:
      - liest Kopfzeilen 'VN:'/'Kunde:'/'Wartung:'/'Anlage:'
      - liest 'Melder:'-Tabelle mit Kopf 'MG' | 'Anz.' | 'Art' | 1..N
      - erkennt wiederholte 'MG'-Kopfzeilen weiter unten automatisch
    """
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    out: Dict[str, Any] = {}
    first_meta = {}

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        # Meta aus Top-4
        meta = _read_top_meta(rows)
        if not first_meta:
            first_meta = meta.copy()

        gruppen = []
        hardware = []  # momentan nicht in Beispielen, vorbereitend gelassen
        wartung = _wartung_defaults()  # falls nicht im Sheet vorhanden

        # Tabelle(n) parsen
        i = 0
        in_table = False
        while i < len(rows):
            r = rows[i]
            first = _s(r, 0)
            if first == "Melder:":
                in_table = True
                i += 2  # nächste Zeile ist Header, überspringen
                continue
            if first == "MG":
                in_table = True
                i += 1  # Header-Zeile überspringen
                continue
            if in_table:
                if _is_blank_row(r):
                    i += 1; continue
                mg = _s(r, 0)
                if not mg or mg == "MG":
                    i += 1; continue
                anz = _int(r, 1)
                art = _s(r, 2)
                melder_typen = []
                for j in range(anz):
                    val = _s(r, 3 + j)
                    melder_typen.append(val)
                gruppen.append({"Nummer": mg, "Anz": anz, "Art": art, "MelderTypen": melder_typen})
            i += 1

        out[ws.title.strip() if ws.title else "-"] = {
            "Gruppen": gruppen,
            "Hardware": hardware,
            "Wartung": wartung
        }

    wb.close()
    return out, first_meta

def _parse_protokoll_workbook(path: Optional[Path]) -> Dict[str, Dict[tuple, str]]:
    """
    Pro Sheet: Map (gruppe, melder_index) -> Auslösung/Status ("Q1","N.i.o.","", "-")
    gleiche Tabellenform wie Listen, aber Zellen enthalten Auslösungen statt Typen.
    """
    if not path:
        return {}
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    out: Dict[str, Dict[tuple, str]] = {}

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        amap: Dict[tuple, str] = {}
        i = 0
        in_table = False
        while i < len(rows):
            r = rows[i]
            first = _s(r, 0)
            if first == "Melder:":
                in_table = True
                i += 2; continue
            if first == "MG":
                in_table = True
                i += 1; continue
            if in_table:
                if _is_blank_row(r):
                    i += 1; continue
                mg = _s(r, 0)
                if not mg or mg == "MG":
                    i += 1; continue
                anz = _int(r, 1)
                for j in range(anz):
                    val = _s(r, 3 + j)
                    amap[(mg, j+1)] = (val or "-")
            i += 1

        out[ws.title.strip() if ws.title else "-"] = amap

    wb.close()
    return out

def _read_top_meta(rows) -> Dict[str, str]:
    meta = {}
    # erwartet Zeilen:
    # ['VN:', <num>], ['Kunde:', <text>], ['Wartung:', <text>], ['Anlage:', <sheetname>]
    for k in range(0, min(6, len(rows))):
        key = _s(rows[k], 0)
        val = _s(rows[k], 1)
        if key.endswith(":"):
            meta[key[:-1]] = val
    return meta

# ----------------------------- TSV-Packer -----------------------------------

def _pack_tsv(model: Dict[str, Any]) -> bytes:
    meta = model["Data"]["Meta"]
    lines: List[List[str]] = []
    lines.append(["#VERSION","1"])
    lines.append(["#GENERATED", datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")])
    lines.append(["#VN", meta.get("VN","-"), meta.get("Kunde","-"), meta.get("Wartungstyp","-")])

    for anlage in meta.get("Anlagen", []):
        lines.append(["#ANLAGE", anlage.get("name","-")])
        lines.append(["#COLUMNS_GM","gruppe_nummer","gruppe_art","melder_nummer","melder_typ","ausluesung"])
        for g in anlage.get("Gruppen", []):
            gnr = g.get("Nummer","-")
            gart= g.get("Art","-")
            for m in g.get("Melder", []):
                lines.append([
                    str(gnr), str(gart),
                    str(m.get("Nummer","-")),
                    str(m.get("typ","-")),
                    str(m.get("ausluesung","-"))
                ])
        # (Hardware/Wartung Sektionen leicht ergänzbar – aktuell leer/Defaults)

    buf = io.StringIO()
    for line in lines:
        buf.write("\t".join("" if v is None else str(v) for v in line) + "\n")
    return buf.getvalue().encode("utf-8")

# ----------------------------- Meta & Utils ---------------------------------

def _resolve_meta(vn_nr: str, meta_from_listen: Dict[str, str]) -> Tuple[str, str]:
    kunde = ""
    wartung = ""
    if os.path.exists(DATA_JSON):
        with open(DATA_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
        for e in data.get("vn", []):
            if _normalize_vn(e.get("vn_nr","")) == _normalize_vn(vn_nr):
                kunde = e.get("kunde","") or ""
                if e.get("objekte"):
                    wartung = e["objekte"][0].get("intervall","") or ""
                break
    if not kunde:
        kunde = meta_from_listen.get("Kunde","")
    if not wartung:
        wartung = meta_from_listen.get("Wartung","")
    return kunde, wartung

def _s(row, idx) -> str:
    if not row or idx >= len(row) or row[idx] is None: return ""
    return str(row[idx]).strip()

def _int(row, idx) -> int:
    s = _s(row, idx)
    if not s: return 0
    try:
        return int(float(s))
    except Exception:
        return 0

def _is_blank_row(row) -> bool:
    return not row or all(c is None or str(c).strip()=="" for c in row)

def _wartung_defaults() -> List[Dict[str,str]]:
    return [{"typ":"Q1","Durch":"","datum":""},
            {"typ":"Q2","Durch":"","datum":""},
            {"typ":"Q3","Durch":"","datum":""},
            {"typ":"Q4","Durch":"","datum":""}]

def maybe_compress_then_encrypt(
    plain_bytes: bytes,
    encrypt_func,
    key_bytes: bytes,
    cfg=None,
):
    """
    Komprimiert optional (abhängig von cfg) und verschlüsselt dann.
    Rückgabe: (cipher_bytes, compressed_bool, original_size_int)
    """
    # --- sichere Defaults ---
    gzip_enabled = True
    gzip_min_size = 1024
    gzip_level = 5
    max_plain = 10 * 1024 * 1024  # 10 MB

    # --- cfg übersteuert Defaults, wenn vorhanden ---
    try:
        if cfg is not None and "server" in cfg:
            sec = cfg["server"]
            # getboolean/getint werfen nicht, wenn fallback übergeben wird
            gzip_enabled = sec.getboolean("gzip_enabled", fallback=gzip_enabled)
            gzip_min_size = sec.getint("gzip_min_size", fallback=gzip_min_size)
            gzip_level = sec.getint("gzip_level", fallback=gzip_level)
            if gzip_level < 1: gzip_level = 1
            if gzip_level > 9: gzip_level = 9
            max_plain = sec.getint("max_plain_response_size", fallback=max_plain)
    except Exception:
        # Falls die Config kaputt ist: auf Defaults laufen
        pass

    size = len(plain_bytes)
    if size > max_plain:
        raise ValueError(f"Antwort zu groß ({size} Bytes), abgebrochen.")

    compressed = False
    out_bytes = plain_bytes
    
    if gzip_enabled and size >= gzip_min_size:
        out_bytes = gzip.compress(plain_bytes, compresslevel=gzip_level)
        compressed = True

    # Jetzt verschlüsseln (deine encrypt_payload erwartet (bytes, key))
    cipher = encrypt_func(out_bytes, key_bytes)
    return cipher, compressed, size
