
from __future__ import annotations
from typing import Dict, Any, List, Optional
from dataclasses import dataclass
from pathlib import Path
import json
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from envelope import Envelope, Meta, Protokoll, Anlage, Table, make_head, make_grid, EditedBy, now_utc_iso

HYPHEN_SENTINELS = {"-", "--", "---", "-----"}

@dataclass
class TopMeta:
    vn: str
    kunde: str
    intervall: str
    type_marker: str

def load_meta_rules(meta_path: str) -> Dict[str, Any]:
    p = Path(meta_path)
    with p.open("r", encoding="utf-8") as f:
        return json.load(f)

def resolve_p_type_from_marker(rules: Dict[str, Any], marker: str) -> Optional[str]:
    for p_type, conf in (rules.get("protocols", {}) or {}).items():
        if str(conf.get("typeMarker","")).strip().lower() == str(marker).strip().lower():
            return p_type
    return None

def val_resolved(ws: Worksheet, r: int, c: int) -> str:
    v = ws.cell(row=r, column=c).value
    if v is None or str(v).strip() == "":
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
                v = ws.cell(row=rng.min_row, column=rng.min_col).value
                break
    return "" if v is None else str(v).strip()

def read_top_meta(ws: Worksheet) -> TopMeta:
    def v(r, c): return val_resolved(ws, r, c)
    return TopMeta(vn=v(1,2), kunde=v(2,2), intervall=v(3,2), type_marker=v(5,2))

def read_anlage_name(ws: Worksheet) -> str:
    name = val_resolved(ws, 4, 2)  # B4
    if not name:
        # fallback to sheet title
        name = ws.title or "Anlage"
    return name

def find_start_row(ws: Worksheet, startkey: str, col: int = 1, max_search: int = 5000) -> Optional[int]:
    key = str(startkey).strip()
    for r in range(1, max_search+1):
        if val_resolved(ws, r, col) == key:
            return r
    return None

def read_header_auto(ws: Worksheet, start_row: int, start_col: int, header_rows: int) -> List[List[str]]:
    out: List[List[str]] = []
    for rr in range(start_row, start_row + header_rows):
        row_vals: List[str] = []
        c = start_col
        while True:
            v = val_resolved(ws, rr, c)
            if v == "":
                break
            row_vals.append(v)
            c += 1
        out.append(row_vals)
    return out

def read_header_fixed(ws: Worksheet, start_row: int, start_col: int, header_rows: int, header_width: int) -> List[List[str]]:
    out: List[List[str]] = []
    for rr in range(start_row, start_row + header_rows):
        out.append([val_resolved(ws, rr, c) for c in range(start_col, start_col + header_width)])
    return out

def read_body_by_cols(ws: Worksheet, start_row: int, cols_abs: List[int]) -> List[List[str]]:
    rows: List[List[str]] = []
    r = start_row
    while True:
        a = val_resolved(ws, r, 1)  # stop when A is empty
        if a == "":
            break
        vals: List[str] = []
        non_empty = False
        for c in cols_abs:
            s = val_resolved(ws, r, c)
            if s in HYPHEN_SENTINELS:
                s = ""
            vals.append(s)
            if s != "":
                non_empty = True
        if non_empty:
            rows.append(vals)
        r += 1
    return rows

def build_table_strict(ws_list: Worksheet, ws_proto: Optional[Worksheet], section_rules: Dict[str, Any]) -> Optional[Table]:
    header_conf = (section_rules or {}).get("header") or {}
    header_rows = int(header_conf.get("rows", 1))
    startkey = str(header_conf.get("startkey", "")).strip()
    spans = header_conf.get("spans", []) or []
    if not startkey:
        return None

    start_row = find_start_row(ws_list, startkey, col=1)
    if start_row is None:
        return None

    # Determine header width: explicit or auto
    layout = (section_rules or {}).get("layout") or {}
    header_start_col = int(layout.get("headerStartCol", 1))
    header_width = layout.get("headerWidth", layout.get("headerLength"))
    if header_width is not None:
        header_width = int(header_width)
    else:
        auto_header = read_header_auto(ws_list, start_row, header_start_col, header_rows)
        header_width = max((len(r) for r in auto_header), default=1)

    # Header only uses headerWidth
    header_rows_data = read_header_fixed(ws_list, start_row, header_start_col, header_rows, header_width)

    # Strict body if descriptorCols or quarterCols given; else fallback using qStartCol or columnsEditable
    desc_cols = layout.get("descriptorCols")
    q_cols = layout.get("quarterCols")
    strict_body = (desc_cols is not None) or (q_cols is not None)

    if strict_body:
        desc_cols = sorted(int(i) for i in (desc_cols or []))
        q_cols = sorted(int(i) for i in (q_cols or []))
        cols_rel = desc_cols + q_cols
        cols_abs = [header_start_col + i for i in cols_rel]

        t_rows = read_body_by_cols(ws_list, start_row + header_rows, cols_abs)
        v_rows: List[List[str]] = []
        if ws_proto is not None:
            proto_start = find_start_row(ws_proto, startkey, col=1)
            if proto_start is not None:
                v_rows = read_body_by_cols(ws_proto, proto_start + header_rows, cols_abs)

        out_rows: List[List[Any]] = []
        max_len = max(len(t_rows), len(v_rows), 0)
        for i in range(max_len):
            tr = t_rows[i] if i < len(t_rows) else []
            vr = v_rows[i] if i < len(v_rows) else []
            row: List[Any] = []
            # descriptors (T)
            for idx in range(len(desc_cols)):
                t = tr[idx] if idx < len(tr) else ""
                row.append(t)
            # quarters (T,V)
            for qi in range(len(q_cols)):
                t = tr[len(desc_cols) + qi] if (len(desc_cols) + qi) < len(tr) else ""
                v = vr[len(desc_cols) + qi] if (len(desc_cols) + qi) < len(vr) else ""
                row.append((t, v))
            out_rows.append(row)

        q_start = len(desc_cols) if (len(desc_cols) + len(q_cols)) > 0 else None

        # Remap columnsEditable onto projected descriptor indices
        ce_raw = (section_rules.get("columnsEditable") or {})
        ce_norm = {int(k): str(v).lower() for k, v in ce_raw.items()}
        ce_out = {str(i): ce_norm.get(desc_cols[i], "string") for i in range(len(desc_cols))}

        head = make_head(header_rows_data[0], spans=spans)
        if len(header_rows_data) > 1:
            head.rows = header_rows_data

        return Table(
            head=head,
            grid=make_grid(out_rows, {int(k): v for k, v in ce_out.items()}, q_start),
            itemsEditable=str(section_rules.get("itemsEditable","no")).lower() == "yes"
        )

    # Non-strict: qStartCol or columnsEditable
    q_start_col = section_rules.get("qStartCol")
    if q_start_col is not None:
        q_start_col = int(q_start_col)
        body_width = header_width
        cols_abs = [header_start_col + i for i in range(body_width)]
        t_rows = read_body_by_cols(ws_list, start_row + header_rows, cols_abs)
        v_rows: List[List[str]] = []
        if ws_proto is not None:
            proto_start = find_start_row(ws_proto, startkey, col=1)
            if proto_start is not None:
                v_rows = read_body_by_cols(ws_proto, proto_start + header_rows, cols_abs)
        out_rows: List[List[Any]] = []
        max_len = max(len(t_rows), len(v_rows), 0)
        for i in range(max_len):
            tr = t_rows[i] if i < len(t_rows) else []
            vr = v_rows[i] if i < len(v_rows) else []
            row: List[Any] = []
            for c in range(body_width):
                t = tr[c] if c < len(tr) else ""
                v = vr[c] if c < len(vr) else ""
                if c >= q_start_col:
                    row.append((t, v))
                else:
                    row.append(t)
            out_rows.append(row)

        ce_raw = (section_rules.get("columnsEditable") or {})
        ce_norm = {int(k): str(v).lower() for k, v in ce_raw.items()}
        head = make_head(header_rows_data[0], spans=spans)
        if len(header_rows_data) > 1:
            head.rows = header_rows_data
        return Table(
            head=head,
            grid=make_grid(out_rows, ce_norm, q_start_col),
            itemsEditable=str(section_rules.get("itemsEditable","no")).lower() == "yes"
        )

    # Fallback: infer qStart from columnsEditable; body_width = headerWidth
    ce_raw = (section_rules.get("columnsEditable") or {})
    ce_norm = {int(k): str(v).lower() for k, v in ce_raw.items()}
    q_start = max(ce_norm.keys()) + 1 if ce_norm else None
    body_width = header_width
    cols_abs = [header_start_col + i for i in range(body_width)]
    t_rows = read_body_by_cols(ws_list, start_row + header_rows, cols_abs)
    v_rows: List[List[str]] = []
    if ws_proto is not None:
        proto_start = find_start_row(ws_proto, startkey, col=1)
        if proto_start is not None:
            v_rows = read_body_by_cols(ws_proto, proto_start + header_rows, cols_abs)
    out_rows: List[List[Any]] = []
    max_len = max(len(t_rows), len(v_rows), 0)
    for i in range(max_len):
        tr = t_rows[i] if i < len(t_rows) else []
        vr = v_rows[i] if i < len(v_rows) else []
        row: List[Any] = []
        for c in range(body_width):
            t = tr[c] if c < len(tr) else ""
            v = vr[c] if c < len(vr) else ""
            if (q_start is not None) and (c >= q_start):
                row.append((t, v))
            else:
                row.append(t)
        out_rows.append(row)
    head = make_head(header_rows_data[0], spans=spans)
    if len(header_rows_data) > 1:
        head.rows = header_rows_data
    return Table(
        head=head,
        grid=make_grid(out_rows, ce_norm, q_start),
        itemsEditable=str(section_rules.get("itemsEditable","no")).lower() == "yes"
    )

def compose_envelope_from_workbooks(list_xlsx: str, prot_xlsx: Optional[str], meta_json: str) -> bytes:
    rules = load_meta_rules(meta_json)

    wb_list = load_workbook(list_xlsx, data_only=True)
    sheets_list = wb_list.sheetnames

    wb_proto = None
    sheets_proto = []
    if prot_xlsx:
        wb_proto = load_workbook(prot_xlsx, data_only=True)
        sheets_proto = wb_proto.sheetnames

    # Meta from first sheet
    ws0 = wb_list[sheets_list[0]]
    top0 = read_top_meta(ws0)
    p_type = resolve_p_type_from_marker(rules, top0.type_marker) or "BMZ"
    p_conf = (rules.get("protocols") or {}).get(p_type, {})

    anlagen = []
    melder_types_accum = set()

    for idx, sname in enumerate(sheets_list):
        wsL = wb_list[sname]
        wsP = None
        if wb_proto:
            if sname in sheets_proto:
                wsP = wb_proto[sname]
            elif idx < len(sheets_proto):
                wsP = wb_proto[sheets_proto[idx]]

        melder = build_table_strict(wsL, wsP, p_conf.get("melder", {}))
        hardware = build_table_strict(wsL, wsP, p_conf.get("hardware", {}))

        if melder is None and hardware is None:
            # nothing to add for this sheet
            continue

        # Anlage name from B4 (or sheet title)
        anlage_name = read_anlage_name(wsL)

        if melder:
            # derive melder types from last descriptor col (qStartCol-1)
            qstart = melder.grid.qStartCol if melder.grid.qStartCol is not None else max((int(k) for k in melder.grid.columnsEditable.keys()), default=0)
            typ_idx = qstart - 1 if qstart > 0 else 0
            for row in melder.grid.body:
                if len(row) > typ_idx:
                    val = str(row[typ_idx].v).strip()
                    if val:
                        melder_types_accum.add(val)

        # melder is required to build a meaningful Anlage, but let's include as long as one table exists
        # If melder is None, we can still include hardware-only Anlage if that fits your model
        if melder is None:
            # Minimal placeholder: empty 0x0 table? Skip for now to keep model clean
            continue

        anlagen.append(Anlage(name=anlage_name, melder=melder, hardware=hardware))

    env = Envelope(
        Meta = Meta(pType=p_type, wType=(top0.intervall or "-"), VNnr=(top0.vn or "-"), Kunde=(top0.kunde or "-")),
        Protokoll = Protokoll(
            melderTypes = sorted(melder_types_accum) or ["I","R","DM"],
            anlagen = anlagen,
            editedBy = EditedBy(name="ProtokollServer", ts=now_utc_iso())
        )
    )
    return env.to_json_bytes()
