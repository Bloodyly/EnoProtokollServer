import os
import json
import logging
from datetime import datetime
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

LISTEN_DIR = "/app/shared/Expose/Listen"
PROTOKOLL_DIR = "/app/shared/Expose/Protokolle"
DATA_PATH = "/app/shared/data.json"

def ensure_data_json():
    if not os.path.exists(DATA_PATH):
        print("[INFO] data.json nicht gefunden – neue Datei wird erstellt.")
        with open(DATA_PATH, "w") as f:
            json.dump({"vn": {}}, f, indent=2)
    else:
        print("[INFO] data.json vorhanden.")

def refresh_data():
    data = {"vn": []}

    if not os.path.exists(LISTEN_DIR):
        print(f"[WARN] Verzeichnis {LISTEN_DIR} nicht gefunden.")
        logger.warning("listenverzeichniss nicht gefunden")
        return

    for filename in os.listdir(LISTEN_DIR):
        if not filename.endswith(".xlsx"):
            continue
        logger.warning("file found")
        file_path = os.path.join(LISTEN_DIR, filename)
        wb = load_workbook(file_path)

        # 🧠 Erste Seite analysieren für Meta-Daten
        sheet = wb[wb.sheetnames[0]]
        try:
            vn_nr = str(sheet["B1"].value).strip()
            kunde = str(sheet["B2"].value).strip()
        except Exception as e:
            print(f"[WARN] Fehler beim Lesen von {filename}: {e}")
            logger.warning("Fehler beim lesen")
            continue

        if not vn_nr or not kunde:
            print(f"[WARN] Datei {filename} enthält keine gültigen Meta-Informationen.")
            logger.warning("datei enthält keine metadaten")
            continue

        # 📄 Alle Objekte (Sheets) analysieren
        objekte = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            objekt_name = str(sheet["B4"].value).strip() if sheet["B4"].value else sheet_name.strip()
            intervall = str(sheet["B3"].value).strip() if sheet["B3"].value else "Quartal"
            meldegruppen = 0
            melder_summe = 0
            melder_ausgelöst = 0
            melder_nio = 0
            wartung_durchfuhrender = ["","","",""]
            wartung_datum = ["","","",""]
            
            for row in sheet.iter_rows(min_row=8, min_col=2, max_col=2):
                zelle = row[0].value
                if zelle is None or str(zelle).strip() == "":
                    break
                try:
                    melder_summe += int(zelle)
                    meldegruppen += 1
                except ValueError:
                    continue

            objekte.append({
                "name": objekt_name,
                "melder_gruppen":meldegruppen,
                "melder_anzahl": melder_summe,
                "melder_ausgeloest": 0,         #wird in 2. funktion aus anderer datei befüllt
                "melder_nio": 0,                #wird in 2. funktion aus anderer datei befüllt
                "intervall": intervall,
                "status": "",               #wird in 2. funktion aus anderer datei angepasst
                "wartung_durchfuhrender":wartung_durchfuhrender, #wird in 2. funktion aus anderer datei befüllt
                "wartung_datum":wartung_datum #wird in 2. funktion aus anderer datei befüllt
            })

        # 🧾 In Struktur eintragen
        data["vn"].append({
            "vn_nr": vn_nr,
            "kunde": kunde,
            "filename" : filename,
            "letzte_bearbeitung": datetime.now().isoformat(),
            "objekte": objekte
        })

    # 💾 Speichern
    with open(DATA_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"[INFO] {len(data['vn'])} Verträge wurden eingelesen und gespeichert.")
    logger.warning("verträge eingelesen")
    
def update_from_protokolle():
    if not os.path.exists(DATA_PATH):
        print("[WARN] Keine data.json vorhanden – Abbruch.")
        return

    with open(DATA_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    updated = 0

    for vn in data["vn"]:
        protokoll_path = os.path.join(PROTOKOLL_DIR, vn["filename"])
        if not os.path.exists(protokoll_path):
            print(f"[INFO] Kein Protokoll für {vn['vn_nr']} vorhanden.")
            continue

        try:
            wb = load_workbook(protokoll_path)
        except Exception as e:
            print(f"[ERROR] Protokoll {vn['filename']} konnte nicht geladen werden: {e}")
            continue

        for obj in vn["objekte"]:
            if obj["name"] not in wb.sheetnames:
                print(f"[WARN] Objekt {obj['name']} nicht in Protokoll {vn['filename']} gefunden.")
                continue

            ws = wb[obj["name"]]

            # 🔢 Zähle ausgelöste Melder (ab Spalte D, Zeile 8+)
            ausgelöst = 0
            nio = 0
            for row in ws.iter_rows(min_row=8):
                for cell in row[3:]:  # Spalte D (Index 3)
                    value = str(cell.value).strip() if cell.value else ""
                    if value in ("Q1", "Q2", "Q3", "Q4", "H1", "H2", "i.O."):
                        ausgelöst += 1
                    elif value == "n.i.O.":
                        nio += 1

            obj["melder_ausgeloest"] = ausgelöst
            obj["melder_nio"] = nio

            # 📅 Wartungsinfos suchen (Zelle mit Inhalt "Wartung Ausgeführt:")
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row[0] and "Wartung Ausgeführt" in str(row[0]):
                    info_start = i + 1
                    break
            else:
                info_start = None

            durchfuehrende = ["", "", "", ""]
            daten = ["", "", "", ""]
            if info_start:
                for idx, row in enumerate(ws.iter_rows(min_row=info_start, max_row=info_start+3), start=0):
                    durchfuehrende[idx] = str(row[3]).strip() if row[3] else ""
                    daten[idx] = str(row[13]).strip() if len(row) > 13 and row[13] else ""

            obj["wartung_durchfuhrender"] = durchfuehrende
            obj["wartung_datum"] = daten

            # 📌 Status aktualisieren
            heute = datetime.now()
            quartal = f"Q{((heute.month - 1) // 3) + 1}"
            halbjahr = "H1" if heute.month <= 6 else "H2"

            if obj["intervall"] == "Quartal":
                status_prefix = quartal
            elif obj["intervall"] == "Halbjahr":
                status_prefix = halbjahr
            else:
                status_prefix = heute.strftime("%Y")

            status_suffix = "offen"
            for datum in daten:
                if datum:
                    status_suffix = "erledigt"
                    break

            obj["status"] = f"{status_prefix} {status_suffix}"

        updated += 1

    with open(DATA_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"[INFO] {updated} VN-Einträge mit Protokollen aktualisiert.")

